from flask import Flask, render_template, flash, redirect, url_for, request
# import MySQL
import pymysql
import secrets
# from flask_mysqldb import MySQL
from wtforms import Form, validators, StringField, FloatField, IntegerField, DateField, SelectField
from wtforms.validators import DataRequired
from datetime import datetime
import MySQLdb
import urllib
import requests
from flask import send_file
from io import BytesIO
from openpyxl.utils import get_column_letter


import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import sqlite3
from flask import Flask, request, jsonify
from datetime import datetime, timedelta

# Create instance of flask app
app = Flask(__name__)

# MySQL Configuration
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_COLLECTION'] = "Shankar"
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = '123456789'
app.config['MYSQL_PORT'] = 3306
app.config['MYSQL_DB'] = 'librarydb'
app.config['MYSQL_CURSORCLASS'] = 'DictCursor'

# Initialise MYSQL
# mysql = MySQL(app)

def get_db_connection():
    connection = pymysql.connect(
        host=app.config['MYSQL_HOST'],
        user=app.config['MYSQL_USER'],
        password=app.config['MYSQL_PASSWORD'],
        port=app.config['MYSQL_PORT'],
        database=app.config['MYSQL_DB'],
        cursorclass=pymysql.cursors.DictCursor
    )
    return connection

###########################################################################

from flask import Flask, render_template, flash, redirect, url_for, request, session
from functools import wraps

# Admin credentials (you can store these securely in environment variables or a database)
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "password123"

# Login Route
@app.route('/', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session['admin_logged_in'] = True
            flash("Login successful!", "success")
            return redirect(url_for('index'))
        else:
            flash("Invalid username or password", "danger")
            return redirect(url_for('login'))

    return render_template('login.html')

# Logout Route
@app.route('/logout')
def logout():
    session.pop('admin_logged_in', None)
    flash("You have been logged out", "success")
    return redirect(url_for('login'))

# Authentication Decorator
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('admin_logged_in'):
            flash("You need to log in as admin to access this page", "danger")
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Protect Routes
@app.route('/admin_dashboard')
@admin_required
def admin_dashboard():
    return render_template('admin_dashboard.html')

# Add this to the top of the file to import session and wraps
from flask import session
from functools import wraps



###########################################################################



# Homepage
@app.route('/home')
@admin_required
def index():
    return render_template('home.html')


# Members
@app.route('/members')
@admin_required
def members():
    # Create MySQLCursor
    connection = get_db_connection()
    cur = connection.cursor()

    # Execute SQL Query
    result = cur.execute("SELECT * FROM members")
    members = cur.fetchall()

    # Render Template
    if result > 0:
        return render_template('members.html', members=members)
    else:
        msg = 'No Members Found'
        return render_template('members.html', warning=msg)

    # Close DB Connection
    cur.close()


# View Details of Member by ID
@app.route('/member/<string:id>')
def viewMember(id):
    # Create MySQLCursor
    # cur = mysql.connection.cursor()
    connection = get_db_connection()
    cur = connection.cursor()

    # Execute SQL Query
    result = cur.execute("SELECT * FROM members WHERE id=%s", [id])
    member = cur.fetchone()

    # Render Template
    if result > 0:
        return render_template('view_member_details.html', member=member)
    else:
        msg = 'This Member Does Not Exist'
        return render_template('view_member_details.html', warning=msg)

    # Close DB Connection
    cur.close()


# Define Add-Member-Form
class AddMember(Form):
    name = StringField('Name', [validators.Length(min=1, max=50)])
    email = StringField('Email', [validators.length(min=6, max=50)])


# Add Member
@app.route('/add_member', methods=['GET', 'POST'])
def add_member():
    # Get form data from request
    form = AddMember(request.form)

    # To handle POST request to route
    if request.method == 'POST' and form.validate():
        name = form.name.data
        email = form.email.data

        # Create MySQLCursor
        # cur = mysql.connection.cursor()
        connection = get_db_connection()
        cur = connection.cursor()

        try:
        # Execute SQL Query
            cur.execute(
                "INSERT INTO members (name, email) VALUES (%s, %s)", (name, email))

            # Commit to DB
            connection.commit()

            # Close DB Connection
            cur.close()

            # Flash Success Message
            flash("New Member Added", "success")

            # Redirect to show all members
            return redirect(url_for('members'))
        
        except Exception as e:
            # Handle duplicate email error
            flash("Error: A member with this name or email already exists.", "danger")
            connection.rollback()
    # To handle GET request to route
    return render_template('add_member.html', form=form)


from werkzeug.utils import secure_filename
import os
import pandas as pd
from datetime import datetime
from app import get_db_connection
# Allowed Excel extensions
ALLOWED_EXTENSIONS = {'xlsx'}

# Check for allowed file type
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Route to insert members using Excel file
@app.route('/insert_members', methods=['GET', 'POST'])
def insert_members():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file part', 'danger')
            return redirect(request.url)

        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'danger')
            return redirect(request.url)

        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join('uploads', filename)
            os.makedirs('uploads', exist_ok=True)
            file.save(filepath)

            try:
                df = pd.read_excel(filepath, engine='openpyxl')
            except Exception as e:
                flash(f'Error reading Excel file: {str(e)}', 'danger')
                return redirect(request.url)

            required_columns = {'name', 'email'}
            df.columns = df.columns.str.lower()

            if not required_columns.issubset(df.columns):
                flash('Excel file must contain columns: name, email', 'danger')
                return redirect(request.url)

            connection = get_db_connection()
            cur = connection.cursor()
            inserted_count = 0

            for _, row in df.iterrows():
                try:
                    cur.execute(
                        "INSERT INTO members (name, email) VALUES (%s, %s)",
                        (row['name'], row['email'])
                    )
                    inserted_count += 1
                except Exception as e:
                    print(f"Error inserting member: {e}")
                    continue

            connection.commit()
            cur.close()
            flash(f'{inserted_count} members successfully inserted.', 'success')
            return redirect(url_for('members'))  # Update to your actual members page

        else:
            flash('Invalid file type. Please upload a .xlsx file.', 'danger')
            return redirect(request.url)

    return render_template('insert_members.html')

@app.route('/export_members', methods=['GET'])
def export_members():
    try:
        connection = get_db_connection()
        cursor = connection.cursor()
        # Fetch all members from the database
        query = "SELECT * FROM members"
        cursor.execute(query)
        members = cursor.fetchall()
        cursor.close()
        connection.close()

        if not members:
            flash("No members found to export", "warning")
            return redirect(url_for('members'))

        # Convert to DataFrame
        df = pd.DataFrame(members)
        
        # Create a BytesIO object to store the Excel file
        output = BytesIO()
        
        # Create a timestamp for the filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Create Excel writer and write data
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Members')
            
            # Auto-adjust columns' width
            worksheet = writer.sheets['Members']
            for i, col in enumerate(df.columns):
                column_width = max(df[col].astype(str).map(len).max(), len(col)) + 2
                worksheet.column_dimensions[get_column_letter(i+1)].width = column_width
        
        # Make sure to get all the data written to the BytesIO object
        output.seek(0)
        
        # Create the file name
        filename = f"library_members_export_{timestamp}.xlsx"
        
        # Return the file for download
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    except Exception as e:
        flash(f"Failed to export members: {str(e)}", "danger")
        return redirect(url_for('members'))

# Edit Member by ID
@app.route('/edit_member/<string:id>', methods=['GET', 'POST'])
def edit_member(id):
    # Get form data from request
    form = AddMember(request.form)

    # To handle POST request to route
    if request.method == 'POST' and form.validate():
        name = form.name.data
        email = form.email.data

        # Create MySQLCursor
        # cur = mysql.connection.cursor()
        connection = get_db_connection()
        cur = connection.cursor()


        # Execute SQL Query
        cur.execute(
            "UPDATE members SET name=%s, email=%s WHERE id=%s", (name, email, id))

        # Commit to DB
        connection.commit()

        # Close DB Connection
        cur.close()

        # Flash Success Message
        flash("Member Updated", "success")

        # Redirect to show all members
        return redirect(url_for('members'))

    # To handle GET request to route

    # To get existing field values of selected member
    # cur2 = mysql.connection.cursor()
    connection = get_db_connection()
    cur2 = connection.cursor()

    result = cur2.execute("SELECT name,email FROM members WHERE id=%s", [id])
    member = cur2.fetchone()
    # To render edit member form
    return render_template('edit_member.html', form=form, member=member)


# Delete Member by ID
# Using POST instead of DELETE because HTML form can only send GET and POST requests
@app.route('/delete_member/<string:id>', methods=['POST'])
def delete_member(id):

    # Create MySQLCursor
    # cur = mysql.connection.cursor()
    connection = get_db_connection()
    cur = connection.cursor()

    # Since deleting parent row can cause a foreign key constraint to fail
    try:
        # Execute SQL Query
        cur.execute("DELETE FROM members WHERE id=%s", [id])

        # Commit to DB
        connection.commit()
    except (MySQLdb.Error, MySQLdb.Warning) as e:
        print(e)
        # Flash Failure Message
        flash("Member could not be deleted", "danger")
        flash(str(e), "danger")

        # Redirect to show all members
        return redirect(url_for('members'))
    finally:
        # Close DB Connection
        cur.close()

    # Flash Success Message
    flash("Member Deleted", "success")

    # Redirect to show all members
    return redirect(url_for('members'))


# Books
@app.route('/books')
@admin_required
def books():
    # Create MySQLCursor
    # cur = mysql.connection.cursor()
    connection = get_db_connection()
    cur = connection.cursor()

    # Execute SQL Query
    result = cur.execute(
        "SELECT id,title,author,total_quantity,available_quantity,rented_count,genre FROM books")
    books = cur.fetchall()
    # Render Template
    if result > 0:
        return render_template('books.html', books=books)
    else:
        msg = 'No Books Found'
        return render_template('books.html', warning=msg)

    # Close DB Connection
    cur.close()


# View Details of Book by ID
@app.route('/book/<string:id>')
def viewBook(id):
    # Create MySQLCursor
    # cur = mysql.connection.cursor()
    connection = get_db_connection()
    cur = connection.cursor()

    # Execute SQL Query
    result = cur.execute("SELECT * FROM books WHERE id=%s", [id])
    book = cur.fetchone()

    # Render Template
    if result > 0:
        return render_template('view_book_details.html', book=book)
    else:
        msg = 'This Book Does Not Exist'
        return render_template('view_book_details.html', warning=msg)

    # Close DB Connection
    cur.close()


# Define Add-Book-Form
class AddBook(Form):
    id = StringField('Book ID', [validators.Length(min=1, max=11)])
    title = StringField('Title', [validators.Length(min=2, max=255), DataRequired()])
    author = StringField('Author(s)', [validators.Length(min=2, max=255), DataRequired()])
    average_rating = FloatField(
        'Average Rating', [validators.Optional(), validators.NumberRange(min=0, max=5)])
    isbn = StringField('ISBN', [validators.Length(min=10, max=10), DataRequired()])
    isbn13 = StringField('ISBN13', [validators.Length(min=13, max=13), DataRequired()])
    language_code = StringField('Language', [validators.Optional(), validators.Length(min=1, max=30)])
    num_pages = IntegerField('No. of Pages', [validators.Optional(), validators.NumberRange(min=1)])
    ratings_count = IntegerField(
        'No. of Ratings', [validators.Optional(), validators.NumberRange(min=0)])
    text_reviews_count = IntegerField(
        'No. of Text Reviews', [validators.Optional(), validators.NumberRange(min=0)])
    publication_date = DateField(
        'Publication Date', [validators.InputRequired(), DataRequired()])
    publisher = StringField('Publisher', [validators.Length(min=2, max=255), DataRequired()])
    total_quantity = IntegerField(
        'Total No. of Books', [validators.NumberRange(min=1), DataRequired()])
    genre = StringField('Genre', [validators.Optional(), validators.Length(min=2, max=255)])


# Add Book
@app.route('/add_book', methods=['GET', 'POST'])
@admin_required
def add_book():
    # Get form data from request
    form = AddBook(request.form)

    # To handle POST request to route
    if request.method == 'POST' and form.validate():

        # Create MySQLCursor
        # cur = mysql.connection.cursor()
        connection = get_db_connection()
        cur = connection.cursor()

        # Check if book with same ID already exists
        result = cur.execute(
            "SELECT id FROM books WHERE id=%s", [form.id.data])
        book = cur.fetchone()
        if(book):
            error = 'Book with that ID already exists'
            return render_template('add_book.html', form=form, error=error)

        # Execute SQL Query
        cur.execute("INSERT INTO books (id,title,author,average_rating,isbn,isbn13,language_code,num_pages,ratings_count,text_reviews_count,publication_date,publisher,genre,total_quantity,available_quantity,rented_count) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)", [
            form.id.data,
            form.title.data,
            form.author.data,
            form.average_rating.data,
            form.isbn.data,
            form.isbn13.data,
            form.language_code.data,
            form.num_pages.data,
            form.ratings_count.data,
            form.text_reviews_count.data,
            form.publication_date.data,
            form.publisher.data,
            form.genre.data,
            form.total_quantity.data,
            # When a book is first added, available_quantity = total_quantity
            form.total_quantity.data,
            form.total_quantity.data
        ])

        # Commit to DB
        connection.commit()

        # Close DB Connection
        cur.close()

        # Flash Success Message
        flash("New Book Added", "success")

        # Redirect to show all books
        return redirect(url_for('books'))

    # To handle GET request to route
    return render_template('add_book.html', form=form)


# from flask import Flask, request, jsonify, render_template, redirect, url_for, flash
from werkzeug.utils import secure_filename
import pandas as pd
import os
from app import get_db_connection

ALLOWED_EXTENSIONS = {'xlsx'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/import_books', methods=['GET', 'POST'])
@admin_required
def import_books():
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file part in the request', 'danger')
            return redirect(request.url)

        file = request.files['file']
        if file.filename == '':
            flash('No file selected', 'danger')
            return redirect(request.url)

        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join('uploads', filename)
            os.makedirs('uploads', exist_ok=True)
            file.save(filepath)

            try:
                df = pd.read_excel(filepath, engine='openpyxl')

                required_columns = [
                    'title', 'author', 'isbn', 'isbn13', 'publication_date', 'publisher',
                    'genre', 'total_quantity', 'available_quantity'
                ]
                for col in required_columns:
                    if col not in df.columns:
                        flash(f'Missing required column: {col}', 'danger')
                        return redirect(request.url)

                df['average_rating'] = pd.to_numeric(df.get('average_rating'), errors='coerce')
                df['publication_date'] = pd.to_datetime(df['publication_date'], errors='coerce')
                df['isbn'] = df['isbn'].apply(lambda x: str(int(float(x))) if pd.notnull(x) else None)
                df['isbn13'] = df['isbn13'].apply(lambda x: str(int(float(x))) if pd.notnull(x) else None)
                df = df.where(pd.notnull(df), None)

                connection = get_db_connection()
                cursor = connection.cursor()

                insert_query = """
                    INSERT INTO books (
                        title, author, average_rating, isbn, isbn13, language_code,
                        num_pages, ratings_count, text_reviews_count, publication_date,
                        publisher, genre, total_quantity, available_quantity, rented_count
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """

                skipped_books = []

                for _, row in df.iterrows():
                    try:
                        cursor.execute(insert_query, (
                            row['title'],
                            row['author'],
                            row.get('average_rating'),
                            row['isbn'],
                            row['isbn13'],
                            row.get('language_code'),
                            row.get('num_pages'),
                            row.get('ratings_count'),
                            row.get('text_reviews_count'),
                            row['publication_date'].strftime('%Y-%m-%d') if pd.notnull(row['publication_date']) else None,
                            row['publisher'],
                            row['genre'],
                            row['total_quantity'],
                            row['available_quantity'],
                            row.get('rented_count', 0)
                        ))
                    except Exception as e:
                        skipped_books.append({'isbn': row.get('isbn'), 'error': str(e)})

                connection.commit()
                cursor.close()
                connection.close()

                # flash(f"Books imported successfully. Skipped: {len(skipped_books)}", "success")
                flash(f"Books imported successfully.", "success")
                return redirect(url_for('books'))

            except Exception as e:
                flash(f"Failed to import books: {str(e)}", "danger")
                return redirect(request.url)

        else:
            flash('Invalid file format. Please upload an Excel (.xlsx) file.', 'danger')
            return redirect(request.url)

    return render_template('import_books.html')

@app.route('/export_books', methods=['GET'])
@admin_required
def export_books():
    try:
        # Get database connection
        connection = get_db_connection()
        cursor = connection.cursor()
        
        # Query to get all books from database
        query = """
            SELECT 
                id, title, author, average_rating, isbn, isbn13, language_code,
                num_pages, ratings_count, text_reviews_count, publication_date,
                publisher, genre, total_quantity, available_quantity, rented_count
            FROM books
        """
        cursor.execute(query)
        books = cursor.fetchall()
        cursor.close()
        connection.close()
        
        if not books:
            flash("No books found to export", "warning")
            return redirect(url_for('books'))
        
        # Convert to DataFrame
        df = pd.DataFrame(books)

        # Create a BytesIO object
        output = BytesIO()
        
        # Create Excel writer
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Books')
            
            # Auto-adjust columns' width
            worksheet = writer.sheets['Books']
            for i, col in enumerate(df.columns):
                column_width = max(df[col].astype(str).map(len).max(), len(col)) + 2
                worksheet.column_dimensions[worksheet.cell(row=1, column=i+1).column_letter].width = column_width
        
        # Set the file pointer at the beginning
        output.seek(0)
        
        # Add timestamp to filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"library_books_export_{timestamp}.xlsx"
        
        # Return the Excel file
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        # logging.error(f"Error exporting books: {str(e)}")
        flash(f"Error exporting books: {str(e)}", "danger")
        return redirect(url_for('books'))


# Edit Book by ID
@app.route('/edit_book/<string:id>', methods=['GET', 'POST'])
@admin_required
def edit_book(id):
    # Get form data from request
    form = AddBook(request.form)

    # Create MySQLCursor
    # cur = mysql.connection.cursor()
    connection = get_db_connection()
    cur = connection.cursor()

    # To get existing values of selected book
    result = cur.execute("SELECT * FROM books WHERE id=%s", [id])
    book = cur.fetchone()

    # To handle POST request to route
    if request.method == 'POST' and form.validate():
        # Check if book with same ID already exists (if ID field is being edited)
        if(form.id.data != id):
            result = cur.execute(
                "SELECT id FROM books WHERE id=%s", [form.id.data])
            book = cur.fetchone()
            if(book):
                error = 'Book with that ID already exists'
                return render_template('edit_book.html', form=form, error=error, book=form.data)

        # Calculate new available_quantity (No. of books available to be rented)
        available_quantity = book['available_quantity'] + \
            (form.total_quantity.data - book['total_quantity'])

        # Execute SQL Query
        cur.execute("UPDATE books SET id=%s,title=%s,author=%s,average_rating=%s,isbn=%s,isbn13=%s,language_code=%s,num_pages=%s,ratings_count=%s,text_reviews_count=%s,publication_date=%s,publisher=%s,genre=%s,total_quantity=%s,available_quantity=%s WHERE id=%s", [
            form.id.data,
            form.title.data,
            form.author.data,
            form.average_rating.data,
            form.isbn.data,
            form.isbn13.data,
            form.language_code.data,
            form.num_pages.data,
            form.ratings_count.data,
            form.text_reviews_count.data,
            form.publication_date.data,
            form.publisher.data,
            form.genre.data,
            form.total_quantity.data,
            available_quantity,
            id])

        # Commit to DB
        connection.commit()

        # Close DB Connection
        cur.close()

        # Flash Success Message
        flash("Book Updated", "success")

        # Redirect to show all books
        return redirect(url_for('books'))

    # To handle GET request to route
    # To render edit book form
    return render_template('edit_book.html', form=form, book=book)


# Delete Book by ID
# Using POST instead of DELETE because HTML form can only send GET and POST requests
@app.route('/delete_book/<string:id>', methods=['POST'])
@admin_required
def delete_book(id):
    # Create MySQLCursor
    # cur = mysql.connection.cursor()
    connection = get_db_connection()
    cur = connection.cursor()

    # Since deleting parent row can cause a foreign key constraint to fail
    try:
        # Execute SQL Query
        cur.execute("DELETE FROM books WHERE id=%s", [id])
        reports()           # new_add
        # Commit to DB
        connection.commit()
    except (MySQLdb.Error, MySQLdb.Warning) as e:

        print(e)
        # Flash Failure Message
        flash("Book could not be deleted", "danger")
        flash(str(e), "danger")

        # Redirect to show all members
        return redirect(url_for('books'))
    finally:
        # Close DB Connection
        cur.close()

    # Flash Success Message
    flash("Book Deleted", "success")

    # Redirect to show all books
    return redirect(url_for('books'))


# Transactions
@app.route('/transactions')
@admin_required
def transactions():
    # Create MySQLCursor
    # cur = mysql.connection.cursor()
    connection = get_db_connection()
    cur = connection.cursor()

    # Execute SQL Query
    result = cur.execute("SELECT * FROM transactions")
    transactions = cur.fetchall()

    # To handle empty fields
    for transaction in transactions:
        for key, value in transaction.items():
            if value is None:
                transaction[key] = "-"

    # Render Template
    if result > 0:
        return render_template('transactions.html', transactions=transactions)
    else:
        msg = 'No Transactions Found'
        return render_template('transactions.html', warning=msg)

    # Close DB Connection
    cur.close()

# Update transactions table
def update_transactions():

    # Create MYSQLCursor
    connection = get_db_connection()
    cur = connection.cursor()

    # Get all active transactions
    cur.execute("SELECT id, borrowed_on, per_day_fee FROM transactions WHERE returned_on IS NULL")
    transactions = cur.fetchall()
    # Update the total_charge of each transaction
    for transaction in transactions:
        borrowed_date = transaction["borrowed_on"]
        today_date = datetime.now()
        difference = (today_date - borrowed_date).days + 1
        total_charge = difference * transaction['per_day_fee']
        # Update the total_charge in the database
        cur.execute("UPDATE transactions SET total_charge = %s WHERE id = %s", (total_charge, transaction["id"]))

    # Commit and close the cursor
    connection.commit()
    cur.close()

def update_outstanding_debt():
    connection = get_db_connection()
    cur = connection.cursor()
    cur.execute("update members set outstanding_debt=0")
    cur.execute("select id, member_id, total_charge, amount_paid from transactions")
    transactions = cur.fetchall()
    for transaction in transactions:
        if transaction['total_charge'] > transaction['amount_paid']:
            cur.execute("update members set outstanding_debt=outstanding_debt+%s where id=%s", (transaction['total_charge']-transaction['amount_paid'], transaction['member_id']))
    connection.commit()
    cur.close()


'''
@app.route('/send_alerts')
def send_alerts():
    SMTP_SERVER = "smtp.gmail.com"
    SMTP_PORT = 587
    SENDER_EMAIL = "sikhakollishankar0503@gmail.com"
    SENDER_PASSWORD = "cicj huni rsfi ouaw"

    connection = get_db_connection()
    cur = connection.cursor()

    cur.execute("SELECT members.email, members.name, books.title, transactions.borrowed_on FROM transactions JOIN members ON transactions.member_id = members.id JOIN books ON transactions.book_id = books.id WHERE transactions.id = %s AND transactions.returned_on IS NULL", [transaction_id])

    user_data = cur.fetchone()
    
    if not user_data:
        connection.close()
        return jsonify({"message": "No active transaction found for given transaction ID."}), 404
    
    borrowed_on_date = user_data['borrowed_on']
    
    if (datetime.now() - borrowed_on_date).days <= 10:
        connection.close()
        return jsonify({"message": "Email alert not needed as borrowed duration is within limit."}), 200
    
    connection.close()

    print((datetime.now() - borrowed_on_date).days)
    
    msg = MIMEMultipart()
    msg['From'] = SENDER_EMAIL
    msg['To'] = 'sikhakollishankar16@gmail.com'
    msg['Subject'] = "Library Due Date Reminder"
    email = 'sikhakollishankar0503@gmail.com'
    # print(name,email)
    body = f"""
    Dear '{user_data['name']}',
    
    This is a reminder that you have borrowed the book '{user_data.book}' for more than 10 days. Please return it at the earliest to avoid penalties.
    
    Regards,
    Library Management
    """
    msg.attach(MIMEText(body, 'plain'))
    
    try:
        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(SENDER_EMAIL, SENDER_PASSWORD)
        server.send_message(msg)
        server.quit()
        return jsonify({"message": f"Email sent to {email}"}), 200
    except Exception as e:
        return jsonify({"error": f"Failed to send email: {e}"}), 500
'''

@app.route('/send_email/<string:transaction_id>', methods=['GET'])
@admin_required
def send_email(transaction_id):
    """Send reminder email for overdue books to library members."""
    try:
        # Load email configuration from environment variables for security
        SMTP_SERVER = "smtp.gmail.com"
        SMTP_PORT = 587
        SENDER_EMAIL = "sikhakollishankar0503@gmail.com"
        SENDER_PASSWORD = "cicj huni rsfi ouaw"
        
        # Validate transaction_id to prevent SQL injection
        if not transaction_id.isdigit():
            return jsonify({"error": "Invalid transaction ID format"}), 400
            
        # Get transaction data from database
        connection = get_db_connection()
        cursor = connection.cursor()  # Use dictionary cursor for MySQL
        
        try:
            # Query to get transaction details with book and member information
            cursor.execute("""
                SELECT 
                    m.email, 
                    m.name, 
                    b.title, 
                    t.borrowed_on,
                    t.per_day_fee,
                    DATEDIFF(CURRENT_TIMESTAMP, t.borrowed_on) as days_borrowed
                FROM transactions t
                JOIN members m ON t.member_id = m.id 
                JOIN books b ON t.book_id = b.id 
                WHERE t.id = %s AND t.returned_on IS NULL
            """, (transaction_id,))
            
            user_data = cursor.fetchone()
            
            if not user_data:
                return jsonify({"message": "No active transaction found for given transaction ID."}), 404
            
            # Calculate days borrowed and due date (10 days after borrow date)
            days_borrowed = user_data['days_borrowed']
            borrowed_on_date = user_data['borrowed_on']
            due_date = borrowed_on_date + timedelta(days=10)
            days_overdue = max(0, days_borrowed)
            
            # Calculate current charges
            current_charge = days_overdue * user_data['per_day_fee']
            
            # Don't send email if not overdue
            if days_borrowed <= 7:
                return jsonify({
                    "message": "Email alert not needed as borrowed duration is within limit.",
                    "days_borrowed": days_borrowed
                }), 200
                
            # Log transaction details
            app.logger.info(f"Sending overdue notice for transaction {transaction_id}. Days borrowed: {days_borrowed}")
            
            # Create email
            msg = MIMEMultipart()
            msg['From'] = SENDER_EMAIL
            recipient_email = user_data['email']
            msg['To'] = recipient_email
            msg['Subject'] = f"OVERDUE: Library Book '{user_data['title']}' Reminder"
            
            # Add HTML alternative for better formatting
            html_body = f"""
            <html>
            <body style="font-family: Arial, sans-serif; line-height: 1.6;">
                <p>Dear <b>{user_data['name']}</b>,</p>
                
                <p>This is a reminder that you borrowed the book "<b>{user_data['title']}</b>" on 
                {borrowed_on_date.strftime('%B %d, %Y')} and it is now <span style="color:red;"><b>{days_overdue} days overdue</b></span>.</p>
                
                <p><b>Current charges: ${current_charge:.2f}</b> (${user_data['per_day_fee']:.2f} per day × {days_overdue} days)</p>
                
                <p>Please return it at your earliest convenience to avoid additional late fees.</p>
                
                <h3>Book details:</h3>
                <ul>
                    <li><b>Title:</b> {user_data['title']}</li>
                    <li><b>Borrowed on:</b> {borrowed_on_date.strftime('%B %d, %Y')}</li>
                    <li><b>Due date:</b> {due_date.strftime('%B %d, %Y')}</li>
                    <li><b>Days overdue:</b> {days_overdue}</li>
                </ul>
                
                <p>If you've already returned this book, please disregard this message.</p>
                
                <p>Regards,<br>
                Library Management Team</p>
            </body>
            </html>
            """
            msg.attach(MIMEText(html_body, 'html'))
            
            # Send email with proper error handling
            with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
                server.starttls()
                server.login(SENDER_EMAIL, SENDER_PASSWORD)
                server.send_message(msg)
            
            # Check if email_notifications table exists, create if it doesn't
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS `email_notifications` (
                    `id` INT(11) NOT NULL AUTO_INCREMENT,
                    `transaction_id` INT(11) NOT NULL,
                    `sent_on` TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    `days_overdue` INT(11) NOT NULL,
                    PRIMARY KEY (`id`),
                    FOREIGN KEY (`transaction_id`) REFERENCES transactions(`id`) ON DELETE CASCADE
                ) ENGINE = InnoDB;
            """)
            connection.commit()
                
            # Record email notification in database
            cursor.execute(
                "INSERT INTO email_notifications (transaction_id, sent_on, days_overdue) VALUES (%s, %s, %s)",
                (transaction_id, datetime.now(), days_overdue)
            )
            connection.commit()
                
            return jsonify({
                "success": True, 
                "message": f"Reminder email sent to {recipient_email}",
                "days_overdue": days_overdue,
                "current_charge": current_charge
            }), 200
            
        except Exception as db_error:
            connection.rollback()
            app.logger.error(f"Database error: {str(db_error)}")
            return jsonify({"error": f"Database error: {str(db_error)}"}), 500
        finally:
            if cursor:
                cursor.close()
            if connection:
                connection.close()
                
    except smtplib.SMTPException as smtp_error:
        app.logger.error(f"SMTP error: {str(smtp_error)}")
        return jsonify({"error": f"Failed to send email: {str(smtp_error)}"}), 500
    except Exception as e:
        app.logger.error(f"Unexpected error: {str(e)}")
        return jsonify({"error": f"An unexpected error occurred: {str(e)}"}), 500

# Define Issue-Book-Form
class IssueBook(Form):
    book_id = SelectField('Book ID', choices=[])
    member_id = SelectField('Member ID', choices=[])
    per_day_fee = FloatField('Per Day Renting Fee', [
                             validators.NumberRange(min=1)])


@app.route('/issue_book', methods=['GET', 'POST'])
@admin_required
def issue_book():
    form = IssueBook(request.form)
    connection = get_db_connection()
    cur = connection.cursor()

    cur.execute("SELECT id, title FROM books")
    books = cur.fetchall()
    book_ids_list = [(book['id'], book['title']) for book in books]

    cur.execute("SELECT id, name FROM members")
    members = cur.fetchall()
    member_ids_list = [(member['id'], member['name']) for member in members]

    form.book_id.choices = book_ids_list
    form.member_id.choices = member_ids_list

    if request.method == 'POST' and form.validate():
        cur.execute("SELECT available_quantity FROM books WHERE id=%s", [form.book_id.data])
        result = cur.fetchone()
        available_quantity = result['available_quantity']

        if available_quantity < 1:
            error = 'No copies of this book are available to be rented'
            return render_template('issue_book.html', form=form, error=error)

        total_charge = form.per_day_fee.data  # Minimum charge = 1 day

        cur.execute("""
            INSERT INTO transactions (book_id, member_id, per_day_fee, total_charge, amount_paid, borrowed_on)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, [form.book_id.data, form.member_id.data, form.per_day_fee.data, total_charge, 0.0, datetime.now()])

        cur.execute("""
            UPDATE books SET available_quantity = available_quantity - 1,
                             rented_count = rented_count + 1
            WHERE id=%s
        """, [form.book_id.data])

        # Update member's debt
        cur.execute("SELECT outstanding_debt FROM members WHERE id=%s", [form.member_id.data])
        existing_debt = float(cur.fetchone()['outstanding_debt'])
        updated_debt = existing_debt + total_charge

        cur.execute("""
            UPDATE members SET outstanding_debt = %s WHERE id = %s
        """, [updated_debt, form.member_id.data])

        # Get additional data for email
        cur.execute("SELECT title FROM books WHERE id=%s", [form.book_id.data])
        book_title = cur.fetchone()['title']

        cur.execute("SELECT name, email FROM members WHERE id=%s", [form.member_id.data])
        member_info = cur.fetchone()
        member_name = member_info['name']
        member_email = member_info['email']

        connection.commit()
        cur.close()

        # Calculate return date
        issue_date = datetime.now()
        return_date = issue_date + timedelta(days=7)


        SMTP_SERVER = "smtp.gmail.com"
        SMTP_PORT = 587
        SENDER_EMAIL = "sikhakollishankar0503@gmail.com"
        SENDER_PASSWORD = "cicj huni rsfi ouaw"

        msg = MIMEMultipart()
        msg['From'] = SENDER_EMAIL
        recipient_email = member_email
        msg['To'] = recipient_email
        print(recipient_email, member_email)
        msg['Subject'] = "Details for your Transaction"

        # Send email

        html_body = f"""
<!DOCTYPE html>
<html>
<head>
    <style>
        body {{
            font-family: Arial, sans-serif;
            background-color: #f9f9f9;
            color: #333;
            margin: 0;
            padding: 0;
        }}
        .container {{
            width: 100%;
            max-width: 600px;
            margin: 20px auto;
            background-color: #ffffff;
            border-radius: 8px;
            padding: 20px;
            box-shadow: 0 4px 8px rgba(0, 0, 0, 0.1);
        }}
        .header {{
            text-align: center;
            margin-bottom: 20px;
        }}
        .header h2 {{
            color: #4CAF50;
            font-size: 24px;
            margin: 0;
        }}
        .content {{
            font-size: 16px;
            line-height: 1.5;
        }}
        .content p {{
            margin-bottom: 15px;
        }}
        .footer {{
            font-size: 12px;
            text-align: center;
            color: #777;
            margin-top: 30px;
        }}
        .footer a {{
            color: #4CAF50;
            text-decoration: none;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h2>Library Book Issued</h2>
        </div>
        <div class="content">
            <p>Dear {member_name},</p>
            <p>You have been issued the following book from the library:</p>
            <table style="width: 100%; border-spacing: 10px 0;">
                <tr>
                    <td><strong>Book Title:</strong></td>
                    <td>{book_title}</td>
                </tr>
                <tr>
                    <td><strong>Issued On:</strong></td>
                    <td>{issue_date.strftime('%Y-%m-%d')}</td>
                </tr>
                <tr>
                    <td><strong>Return By:</strong></td>
                    <td>{return_date.strftime('%Y-%m-%d')}</td>
                </tr>
                <tr>
                    <td><strong>Per Day Fee:</strong></td>
                    <td>₹{form.per_day_fee.data}</td>
                </tr>
            </table>
            <p>Please ensure the book is returned on or before the return date to avoid additional charges.</p>
        </div>
        <div class="footer">
            <p>Thank you,<br>Library Management Team</p>
            <p>For any inquiries, contact us at <a href="mailto:library@example.com">library@example.com</a></p>
        </div>
    </div>
</body>
</html>
"""

        try:
            msg.attach(MIMEText(html_body, 'html'))
            with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
                server.starttls()
                server.login(SENDER_EMAIL, SENDER_PASSWORD)
                server.send_message(msg)
        except Exception as email_err:
            app.logger.error(f"Failed to send issue notification: {str(email_err)}")
            print(email_err)
        except Exception as e:
            print(e)

        flash("Book Issued and Email Sent", "success")
        return redirect(url_for('transactions'))

    return render_template('issue_book.html', form=form)



class ReturnBook(Form):
    amount_paid = FloatField('Amount Paid', [validators.NumberRange(min=0)])

# Define Issue-Book-Form
@app.route('/return_book/<string:transaction_id>', methods=['GET', 'POST'])
@admin_required
def return_book(transaction_id):
    form = ReturnBook(request.form)
    connection = get_db_connection()
    cur = connection.cursor()

    cur.execute("SELECT * FROM transactions WHERE id=%s", [transaction_id])
    transaction = cur.fetchone()

    if not transaction:
        flash("Transaction not found", "danger")
        return redirect(url_for('transactions'))

    is_returned = transaction['returned_on'] is not None
    total_charge = transaction['total_charge']
    previous_payment = float(transaction['amount_paid'] or 0)
    remaining_due = total_charge - previous_payment
    date = datetime.now()
    difference = date - transaction['borrowed_on']

    if not is_returned:
        date = datetime.now()
        difference = date - transaction['borrowed_on']
        borrowed_days = difference.days+1
        # borrowed_days = (datetime.now().date() - transaction['borrowed_on']).days + 1
        total_charge = borrowed_days * transaction['per_day_fee']
        remaining_due = total_charge  # Full amount since nothing paid yet

    if request.method == 'POST' and form.validate():
        amount_paid = form.amount_paid.data

        cur.execute("SELECT outstanding_debt, amount_spent FROM members WHERE id=%s", [transaction['member_id']])
        member = cur.fetchone()
        current_debt = float(member['outstanding_debt'])
        current_spent = float(member['amount_spent'])

        if is_returned:
            # Existing return: only process new payment
            new_paid = previous_payment + amount_paid
            cur.execute("UPDATE transactions SET amount_paid=%s WHERE id=%s", [new_paid, transaction_id])

            new_debt = max(0, current_debt - amount_paid)
            cur.execute("UPDATE members SET outstanding_debt=%s, amount_spent=%s WHERE id=%s", [
                new_debt, current_spent + amount_paid, transaction['member_id']
            ])
        else:
            # First time return: finalize transaction
            transaction_debt = max(0, total_charge - amount_paid)

            if current_debt - previous_payment + transaction_debt > 5000:
                error = 'Outstanding Debt Cannot Exceed Rs.500'
                return render_template('return_book.html', form=form, error=error,
                                       payment_to_be_done=total_charge, difference=borrowed_days,
                                       transaction=transaction)

            cur.execute("""
                UPDATE transactions SET returned_on=%s, total_charge=%s, amount_paid=%s WHERE id=%s
            """, [datetime.now(), total_charge, amount_paid, transaction_id])

            new_debt = current_debt - amount_paid
            cur.execute("""
                UPDATE members SET outstanding_debt=%s, amount_spent=%s WHERE id=%s
            """, [new_debt, current_spent + amount_paid, transaction['member_id']])

            # Return the book
            cur.execute("""
                UPDATE books SET available_quantity = available_quantity + 1 WHERE id = %s
            """, [transaction['book_id']])

        connection.commit()
        cur.close()

        flash("Payment Recorded" if is_returned else "Book Returned", "success")
        return redirect(url_for('transactions'))

    # GET request handling
    return render_template('return_book.html', form=form,
                           difference=difference.days+1,
                           total_charge=total_charge,
                           previous_payment=previous_payment,
                           remaining_debt=remaining_due,
                           transaction=transaction,
                           is_returned=is_returned)


@app.route('/transaction/<string:transaction_id>')
@admin_required
def delete_transaction(transaction_id):
    # Create MySQLCursor
    # cur = mysql.connection.cursor()
    connection = get_db_connection()
    cur = connection.cursor()

    # get the transaction record and delete the record
    cur.execute('DELETE FROM transactions where id=%s', [transaction_id])

    # Commit to DB
    connection.commit()

    # Close DB Connection
    cur.close()

    # Flash Success Message
    flash("Book Returned", "success")
    return redirect(url_for('transactions'))
    # return render_template('transactions.html')



@app.route('/reports')
@admin_required
def reports():
    # Get filter parameters with default values
    book_count = request.args.get('book_count', default=5, type=int)
    member_count = request.args.get('member_count', default=5, type=int)
    
    # Ensure valid values (fallback to defaults if invalid)
    # valid_counts = [5, 10, 20, 50]
    # if book_count not in valid_counts:
    #     book_count = 5
    # if member_count not in valid_counts:
    #     member_count = 5
    
    # Create database connection
    connection = get_db_connection()
    cur = connection.cursor()
    
    # Execute SQL Query to get N highest paying members
    result_members = cur.execute(
        "SELECT id, name, amount_spent FROM members ORDER BY amount_spent DESC LIMIT %s",
        (member_count,))
    members = cur.fetchall()
    
    # Execute SQL Query to get N most popular books
    result_books = cur.execute(
        "SELECT id, title, author, total_quantity, available_quantity, rented_count FROM books ORDER BY rented_count DESC LIMIT %s",
        (book_count,))
    books = cur.fetchall()
    
    # Build warning message if no results
    msg = ''
    if result_members <= 0:
        msg = 'No Members Found. '
    if result_books <= 0:
        msg = msg + 'No Books Found'
    
    # Close the cursor
    cur.close()
    
    # Render the template with all necessary data
    return render_template(
        'reports.html', 
        members=members, 
        books=books, 
        warning=msg,
        book_count=book_count,
        member_count=member_count
    )

# Define Search-Form
class SearchBook(Form):
    title = StringField('Title', [validators.Length(min=2, max=255), validators.Optional()])
    author = StringField('Author(s)', [validators.Length(min=2, max=255), validators.Optional()])
    genre = StringField('Genre', [validators.Length(min=2, max=255), validators.Optional()])

# Search
@app.route('/search_book', methods=['GET', 'POST'])
@admin_required
def search_book():
    # Get form data from request
    form = SearchBook(request.form)

    # To handle POST request to route
    if request.method == 'POST' and form.validate():
        # Get search terms
        title = form.title.data.strip()
        author = form.author.data.strip()
        genre = form.genre.data.strip()
        
        # Check if at least one field has data
        if not (title or author or genre):
            msg = 'Please enter at least one field to search'
            return render_template('search_book.html', form=form, warning=msg)
        
        # Create query dynamically based on which fields are filled
        connection = get_db_connection()
        cur = connection.cursor()
        
        query = "SELECT * FROM books WHERE "
        params = []
        conditions = []
        
        if title:
            conditions.append("title LIKE %s")
            params.append('%' + title + '%')
        
        if author:
            conditions.append("author LIKE %s")
            params.append('%' + author + '%')
        
        if genre:
            conditions.append("genre LIKE %s")
            params.append('%' + genre + '%')
        
        # Join conditions with AND for more precise search
        query += " AND ".join(conditions)
        
        # Execute query
        result = cur.execute(query, params)
        books = cur.fetchall()
        cur.close()

        # Flash Success Message
        if result <= 0:
            msg = 'No Results Found'
            return render_template('search_book.html', form=form, warning=msg)

        flash("Results Found", "success")
        # Render template with search results
        return render_template('search_book.html', form=form, books=books)

    # To handle GET request to route
    return render_template('search_book.html', form=form)


@app.route("/digital_books")
@admin_required
def digital_books(): 
    
    # Create MySQLCursor
    # cur = mysql.connection.cursor()
    connection = get_db_connection()
    cur = connection.cursor()

    result = cur.execute("SELECT id, title, author, publisher, cost from digitalbooks")
    books = cur.fetchall()
    # print(books)
    return render_template("digital_books.html",  books=books)
    
# Define add-book form
class AddDigitalBook(Form):
    id = StringField('BookID', [validators.Length(min=1, max=111)])
    title = StringField('Title', [validators.Length(min=2, max=255)])
    author = StringField('Author(s)', [validators.Length(min=2, max=255)])
    publisher = StringField('Publisher(s)', [validators.Length(min=2, max=255)])
    cost = IntegerField('Cost')
    access_link = StringField('access_link', [validators.Length(min=2, max=255)])
    
@app.route("/add_digital_book", methods=['GET','POST'])
@admin_required
def add_digital_book():
    form=AddDigitalBook(request.form)
    if request.method == 'POST' and form.validate():
        # Create MySQLCursor
        # cur = mysql.connection.cursor()
        connection = get_db_connection()
        cur = connection.cursor()

        # Execute SQL Query
        cur.execute("INSERT INTO digitalbooks (id, title, author, publisher, cost, access_link) VALUES (%s,%s,%s,%s,%s,%s)",[
            form.id.data,
            form.title.data,
            form.author.data,
            form.publisher.data,
            form.cost.data,
            form.access_link.data
        ])

        # Commit to DB
        connection.commit()
        
        # Close DB Connection
        cur.close()

        # Flash Success Message
        flash("New Book Added", "success")

        # Redirect to show all books
        return redirect(url_for('digital_books'))
    
    # To handle GET request to route
    return render_template("add_digital_book.html",form=form)


# ------------------------------------------------------

# Form Class
class IssueDigitalBook(Form):
    book_id = SelectField('Book ID', choices=[])
    member_id = SelectField('Member ID', choices=[])

# Function to Generate Secure Access Token
def generate_access_token():
    return secrets.token_urlsafe(32)

# Function to Send Secure Email
def send_email(member_email, book_title, access_token, member_name=None, expiry_date=None, book_author=None, purchase_id=None):
    """
    Send a professional email notification for digital book purchases with detailed information.
    
    Args:
        member_email (str): Recipient's email address
        book_title (str): Title of the purchased digital book
        access_token (str): Unique token for accessing the digital book
        member_name (str, optional): Name of the member who made the purchase
        expiry_date (datetime, optional): When the access token expires
        book_author (str, optional): Author of the purchased book
        purchase_id (int, optional): Purchase transaction ID for reference
    
    Returns:
        tuple: (success_boolean, message_string)
    """
    try:
        # Load email configuration from environment variables for security
        smtp_server = "smtp.gmail.com"
        smtp_port = 587
        sender_email = "sikhakollishankar0503@gmail.com"
        sender_password = "cicj huni rsfi ouaw"
        
        if not sender_password:
            app.logger.error("Email password not configured in environment variables")
            return False, "Email configuration error: missing password"
        
        # Validate inputs
        if not all([member_email, book_title, access_token]):
            app.logger.error("Missing required email parameters")
            return False, "Missing required parameters"
            
        # Set defaults for optional parameters
        member_name = member_name or "Valued Customer"
        expiry_date = expiry_date or (datetime.now() + timedelta(days=7))
        # receipt_number = f"LIB-{purchase_id}" if purchase_id else f"LIB-{int(datetime.time())}"
        
        # Create secure access link with HTTPS for production
        base_url = "http://127.0.0.1:5000"
        access_link = f"{base_url}/view-digital-book/{access_token}"
        
        # Create a multipart message for both plain text and HTML
        msg = MIMEMultipart("alternative")
        msg["Subject"] = f"Your Digital Book Purchase: {book_title}"
        msg["From"] = sender_email
        msg["To"] = member_email
        # msg["Message-ID"] = f"<{uuid.uuid4()}@library.com>"
        # msg["Date"] = formatdate(localtime=True)
        # msg["X-Receipt-Number"] = receipt_number
        
        # Plain text version of the email
        text_body = f"""Hello {member_name},

Thank you for your purchase of '{book_title}'{' by ' + book_author if book_author else ''}!

Your purchase details:
- Book: {book_title}
- Purchase Date: {datetime.now().strftime('%B %d, %Y')}
- Access Expires: {expiry_date.strftime('%B %d, %Y')}

To view your digital book, please visit:
{access_link}

Important Notes:
- This link is personal and unique to your account
- The access link will expire on {expiry_date.strftime('%B %d, %Y')}
- Do not share this link with others as it's connected to your account

If you experience any issues accessing your digital book, please contact our support team.

Thank you for using our digital library services!

Best regards,
Library Management Team
"""
        # HTML version of the email
        html_body = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
</head>
<body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333; max-width: 600px; margin: 0 auto; padding: 20px;">
    <div style="background-color: #f0f5ff; padding: 20px; border-radius: 5px; border-left: 4px solid #3366cc;">
        <h2 style="color: #3366cc; margin-top: 0;">Your Digital Book Purchase</h2>
        <p>Hello <strong>{member_name}</strong>,</p>
        
        <p>Thank you for your purchase of <strong>'{book_title}'</strong>{' by <em>' + book_author + '</em>' if book_author else ''}!</p>
        
        <div style="background-color: white; border-radius: 5px; padding: 15px; margin: 20px 0; border: 1px solid #e0e0e0;">
            <h3 style="margin-top: 0; color: #3366cc;">Purchase Details</h3>
            <table style="width: 100%; border-collapse: collapse;">
                <tr>
                    <td style="padding: 8px 0; border-bottom: 1px solid #eee; width: 40%;"><strong>Book:</strong></td>
                    <td style="padding: 8px 0; border-bottom: 1px solid #eee;">{book_title}</td>
                </tr>
                
                <tr>
                    <td style="padding: 8px 0; border-bottom: 1px solid #eee;"><strong>Purchase Date:</strong></td>
                    <td style="padding: 8px 0; border-bottom: 1px solid #eee;">{datetime.now().strftime('%B %d, %Y')}</td>
                </tr>
                <tr>
                    <td style="padding: 8px 0; border-bottom: 1px solid #eee;"><strong>Access Expires:</strong></td>
                    <td style="padding: 8px 0; border-bottom: 1px solid #eee;">{expiry_date.strftime('%B %d, %Y')}</td>
                </tr>
            </table>
        </div>
        
        <div style="text-align: center; margin: 30px 0;">
            <a href="{access_link}" style="background-color: #3366cc; color: white; padding: 12px 25px; text-decoration: none; border-radius: 4px; font-weight: bold; display: inline-block;">Access Your Digital Book</a>
        </div>
        
        <div style="background-color: #fff4e5; padding: 15px; border-radius: 5px; margin-top: 20px; border-left: 4px solid #ff9900;">
            <h4 style="margin-top: 0; color: #ff9900;">Important Notes:</h4>
            <ul style="padding-left: 20px; margin-bottom: 0;">
                <li>This link is personal and unique to your account</li>
                <li>The access link will expire on <strong>{expiry_date.strftime('%B %d, %Y')}</strong></li>
                <li>Do not share this link with others as it's connected to your account</li>
            </ul>
        </div>
        
        <p style="margin-top: 30px;">If you experience any issues accessing your digital book, please contact our support team.</p>
        
        <p>Thank you for using our digital library services!</p>
        
        <p style="margin-bottom: 0;">Best regards,<br>
        <strong>Library Management Team</strong></p>
    </div>
    <div style="text-align: center; font-size: 12px; color: #777; margin-top: 20px;">
        &copy; {datetime.now().year} Digital Library System. All rights reserved.
    </div>
</body>
</html>
"""
        # Attach both versions to the email
        msg.attach(MIMEText(text_body, "plain"))
        msg.attach(MIMEText(html_body, "html"))
        
        # Connect to the SMTP server and send the email
        try:
            with smtplib.SMTP(smtp_server, smtp_port) as server:
                server.starttls()
                server.login(sender_email, sender_password)
                server.send_message(msg)
                
            # Log successful email delivery
            app.logger.info(f"Digital book access email sent to {member_email} for book '{book_title}'")
            return True, f"Email successfully sent to {member_email}"
            
        except smtplib.SMTPException as smtp_error:
            app.logger.error(f"SMTP error: {str(smtp_error)}")
            return False, f"Failed to send email: {str(smtp_error)}"
            
    except Exception as e:
        app.logger.error(f"Unexpected error sending email: {str(e)}")
        return False, f"An unexpected error occurred: {str(e)}"

# Route to Issue Digital Book
@app.route('/issue_digital_book', methods=['GET', 'POST'])
@admin_required
def issue_digital_book():
    form = IssueDigitalBook(request.form)
    connection = get_db_connection()
    cur = connection.cursor()

    # Populate Book & Member Dropdowns
    cur.execute("SELECT id, title, cost FROM digitalbooks")
    books = cur.fetchall()
    form.book_id.choices = [(book['id'], book['title']) for book in books]

    cur.execute("SELECT id, name, email FROM members")
    members = cur.fetchall()
    form.member_id.choices = [(member['id'], member['name']) for member in members]

    if request.method == 'POST' and form.validate():
        # Get Book Cost & Title
        cur.execute("SELECT title, cost FROM digitalbooks WHERE id = %s", [form.book_id.data])
        book = cur.fetchone()
        if not book:
            flash("Book not found!", "danger")
            return redirect(url_for('issue_digital_book'))

        book_title = book['title']
        cost_value = book['cost']

        # Get Member Email
        cur.execute("SELECT email FROM members WHERE id = %s", [form.member_id.data])
        member = cur.fetchone()
        if not member:
            flash("Member not found!", "danger")
            return redirect(url_for('issue_digital_book'))

        member_email = member['email']

        # Generate Secure Token & Expiry Date
        access_token = generate_access_token()
        expiry_date = datetime.now() + timedelta(days=7)  # 7 days expiry

        # Insert Purchase into Database
        cur.execute(
            "INSERT INTO purchases (member_id, digitalbook_id, purchase_date, total_cost, access_token, expires_at) "
            "VALUES (%s, %s, %s, %s, %s, %s)",
            (form.member_id.data, form.book_id.data, datetime.now(), cost_value, access_token, expiry_date)
        )

        # Commit Changes
        connection.commit()
        cur.close()
        connection.close()

        # Send Secure Access Link via Email
        send_email(member_email, book_title, access_token)

        flash("Digital Book Issued & Access Link Sent!", "success")
        return redirect(url_for('digital_books_transactions'))

    return render_template('purchase_digital_book.html', form=form)

# Route to Display All Digital Book Transactions
@app.route('/digital_books_transactions')
@admin_required
def digital_books_transactions():
    connection = get_db_connection()
    cur = connection.cursor()
    
    cur.execute("""
        SELECT p.id, m.id AS member_name, d.title AS book_title, p.purchase_date, p.total_cost, p.access_token, p.expires_at
        FROM purchases p
        JOIN members m ON p.member_id = m.id
        JOIN digitalbooks d ON p.digitalbook_id = d.id
    """)
    purchases = cur.fetchall()
    
    cur.close()
    connection.close()

    if purchases:
        return render_template('digital_books_transactions.html', purchases=purchases, now=datetime.now())
    else:
        flash("No purchases found", "warning")
        return render_template('digital_books_transactions.html', purchases=[])
    
app.secret_key = "your_very_secret_key"  # Set a unique secret key
@app.route("/view-digital-book/<token>")
@admin_required
def view_digital_book(token):
    connection = get_db_connection()
    cur = connection.cursor()

    # Validate Token & Check Expiry
    cur.execute("SELECT d.access_link FROM purchases p JOIN digitalbooks d ON p.digitalbook_id = d.id WHERE p.access_token = %s AND p.expires_at > NOW()", (token,))
    book = cur.fetchone()
    
    cur.close()
    connection.close()

    if book:
        return redirect(book["access_link"])  # Redirect user to the Google Drive link
    # Securely serve file
    else:
        return "Invalid or expired access link", 403


# from flask import Flask, render_template, request, redirect, url_for, session, flash
# from functools import wraps
# import pymysql

# Member Login Route

@app.route('/member_login', methods=['GET', 'POST'])
def member_login():
    if request.method == 'POST':
        email = request.form['email']
        connection = get_db_connection()
        cur = connection.cursor()
        cur.execute("SELECT * FROM members WHERE email = %s", (email,))
        member = cur.fetchone()
        cur.close()
        connection.close()

        if member:
            session['member_id'] = member['id']
            session['member_name'] = member['name']
            flash("Logged in successfully", "success")
            return redirect(url_for('member_dashboard'))
        else:
            flash("Email not found. Please try again.", "danger")
            return redirect(url_for('member_login'))
    return render_template('member_login.html')

def member_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('member_id'):
            flash("You need to log in as a member to access this page", "danger")
            return redirect(url_for('member_login'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/member/dashboard')
@member_required
def member_dashboard():
    connection = get_db_connection()
    cur = connection.cursor()

    cur.execute("SELECT * FROM members WHERE id = %s", (session['member_id'],))
    member_info = cur.fetchone()

    cur.execute("SELECT t.*, b.title FROM transactions t JOIN books b ON t.book_id = b.id WHERE t.member_id = %s", (session['member_id'],))
    transactions = cur.fetchall()

    # Fetch physical books (available)
    cur.execute("SELECT * FROM books WHERE available_quantity > 0")
    books = cur.fetchall()

    # Fetch digital books
    cur.execute("SELECT * FROM digitalbooks")
    digitalbooks = cur.fetchall()

    
    cur.close()
    connection.close()

    return render_template("member_dashboard.html", member=member_info, transactions=transactions, books=books, digitalbooks=digitalbooks)

@app.route('/member/edit', methods=['GET', 'POST'])
@member_required
def edit_member_details():
    connection = get_db_connection()
    cur = connection.cursor()

    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']

        cur.execute("UPDATE members SET name=%s, email=%s WHERE id=%s", (name, email, session['member_id']))
        connection.commit()
        flash("Details updated successfully!", "success")
        return redirect(url_for('member_dashboard'))

    cur.execute("SELECT * FROM members WHERE id = %s", (session['member_id'],))
    member_info = cur.fetchone()
    cur.close()
    connection.close()

    return render_template('edit_member.html', member=member_info)

@app.route('/member_logout')
@member_required
def member_logout():
    session.pop('member_id', None)
    session.pop('member_name', None)
    flash("You have been logged out", "success")
    return redirect(url_for('member_login'))






@app.route('/member/search_books', methods=['GET', 'POST'])
@member_required
def search_books():
    results = []
    if request.method == 'POST':
        query = request.form['query']

        connection = get_db_connection()
        cur = connection.cursor()
        cur.execute("""
            SELECT * FROM books 
            WHERE genre LIKE %s OR title LIKE %s OR author LIKE %s
        """, [('%' + query + '%'),('%' + query + '%'),('%' + query + '%')])
        results = cur.fetchall()
        cur.close()

    return render_template('search_books.html', results=results)


if __name__ == '__main__':
    app.secret_key = "secret"
    update_transactions()
    update_outstanding_debt()
    app.run(debug=True)